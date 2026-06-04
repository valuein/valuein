"""
generate_catalog.py
===================
Generates the public-facing data catalog from the canonical standardized-concept
list published in the R2 ``manifest.json``.

Single source of truth
----------------------
The list of canonical ``fact.standard_concept`` values is no longer hardcoded in
this script.  It is sourced directly from the live, publicly readable R2 manifest
(no API token required):

    https://data.valuein.biz/v1/sample/manifest

The manifest carries a top-level ``standard_concepts`` key produced by the
pipeline (``data-pipeline/run_exports.py`` from
``services/accounting/definitions.py``).  Two shapes are supported so the catalog
keeps working across the pipeline contract migration:

  * **Rich shape (current contract)** — ``list[dict]``, each entry carrying::

        standard_concept, level, statement_type, category, definition,
        unit_default, is_flow, bloomberg_equivalent, factset_equivalent,
        gaap_ifrs_comparable

  * **Legacy shape** — ``list[str]`` of concept names only.  Name-only entries
    are surfaced with empty definition/category/statement_type so the catalog
    still lists them; re-run after the next pipeline export to backfill the rich
    fields.

Outputs:
  docs/data_catalog.md       — Human-readable Markdown for analysts and integration partners.
  docs/data_catalog.json     — Machine-readable JSON for SDK metadata and docs sites.
  docs/DATA_CATALOG.xlsx     — Excel workbook for financial analysts (updates sheet
                               "5. Standardized Concepts" and refreshes the generated date
                               on the Overview sheet; all other sheets are preserved).

Run from the repo root:
    uv run python scripts/generate_catalog.py

Override the manifest source for testing:
    VALUEIN_MANIFEST_URL=https://example/manifest uv run python scripts/generate_catalog.py

Re-run whenever STANDARD_DEFINITIONS in data-pipeline/services/accounting/definitions.py
changes AND a fresh pipeline export has been published — this script reflects
exactly what the live manifest exposes, not a local copy.
"""

from __future__ import annotations

import json
import os
import urllib.error
import urllib.request
from datetime import date, datetime, timezone

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Manifest source — the canonical standardized-concept list lives here.
# Publicly readable, no API token required.
# ---------------------------------------------------------------------------

DEFAULT_MANIFEST_URL = "https://data.valuein.biz/v1/sample/manifest"
_MANIFEST_TIMEOUT_SECONDS = 30

_DOCS_DIR = os.path.join(os.path.dirname(__file__), "..", "docs")

# Canonical statement-type ordering for the Markdown sections. Any
# statement_type returned by the manifest that is not listed here is appended
# afterwards in first-seen order; name-only/legacy concepts fall into the
# "Uncategorized" bucket which always sorts last.
_UNCATEGORIZED = "Uncategorized"
_STATEMENT_ORDER = [
    "Income Statement",
    "Balance Sheet",
    "Cash Flow Statement",
]


class CatalogError(RuntimeError):
    """Raised when the catalog cannot be generated (e.g. manifest unavailable
    or has not yet published any standardized concepts)."""


# ---------------------------------------------------------------------------
# Financial-ratio catalog (stable IP — embedded here, NOT sourced from the
# manifest).  The R2 manifest publishes the ``standard_concept`` list (raw
# fact rows) but does not yet carry ratio definitions, so the authoritative
# ratio inventory lives as a Python constant.  Mirrors the ``ratio`` table in
# data-pipeline (services + pipeline/steps/augmentation.py).  Re-derive from
# that table when the pipeline adds/renames a ratio.
#
# Each entry is ``(ratio_name, category, unit, ttm_available, description)``.
# TTM (trailing-twelve-month) rows exist ONLY for the categories flagged
# ``ttm_available=True`` below: profitability, per_share, liquidity,
# efficiency, leverage.  forensic, growth (CAGR), and the sector-percentile
# ranks are annual-only (one snapshot per fiscal year).
# ---------------------------------------------------------------------------

# Categories whose ratios are materialized for BOTH FY and TTM periods.
_TTM_CATEGORIES = frozenset(
    {"profitability", "per_share", "liquidity", "efficiency", "leverage"}
)

# Display order for the rendered Markdown / Excel sections.
_RATIO_CATEGORY_ORDER = [
    "profitability",
    "per_share",
    "liquidity",
    "efficiency",
    "leverage",
    "forensic",
    "growth",
    "rank",
]

# Human-readable headings for each ratio category.
_RATIO_CATEGORY_LABELS = {
    "profitability": "Profitability",
    "per_share": "Per Share",
    "liquidity": "Liquidity",
    "efficiency": "Efficiency",
    "leverage": "Leverage",
    "forensic": "Forensic",
    "growth": "Growth / CAGR",
    "rank": "Sector Percentile Ranks",
}

# --- profitability (32) — TTM available -----------------------------------
_PROFITABILITY_RATIOS: list[tuple[str, str, str]] = [
    ("gross_profit_margin", "ratio", "Gross profit divided by revenue."),
    ("operating_margin", "ratio", "Operating income divided by revenue."),
    ("net_margin", "ratio", "Net income divided by revenue."),
    ("pretax_income_margin", "ratio", "Pre-tax income divided by revenue."),
    ("cost_of_revenue_margin", "ratio", "Cost of revenue divided by revenue."),
    ("opex_margin", "ratio", "Operating expenses divided by revenue."),
    ("da_margin", "ratio", "Depreciation & amortization divided by revenue."),
    ("interest_expense_margin", "ratio", "Interest expense divided by revenue."),
    ("net_interest_margin", "ratio", "Net interest income divided by revenue (financials)."),
    ("sga_margin", "ratio", "Selling, general & administrative expense divided by revenue."),
    ("total_expense_margin", "ratio", "Total expenses divided by revenue."),
    ("rd_margin", "ratio", "Research & development expense divided by revenue."),
    ("ocf_margin", "ratio", "Operating cash flow divided by revenue."),
    ("fcf_margin", "ratio", "Free cash flow divided by revenue."),
    ("roic", "ratio", "Return on invested capital — NOPAT divided by invested capital."),
    ("invested_capital", "USD", "Total invested capital (debt + equity, net of cash)."),
    ("nopat", "USD", "Net operating profit after tax."),
    ("return_on_assets", "ratio", "Net income divided by average total assets."),
    ("return_on_equity", "ratio", "Net income divided by average shareholders' equity."),
    ("ebitda", "USD", "Earnings before interest, taxes, depreciation & amortization."),
    ("ebitda_margin", "ratio", "EBITDA divided by revenue."),
    ("roce", "ratio", "Return on capital employed — EBIT divided by capital employed."),
    ("roce_nopat", "ratio", "Return on capital employed computed on NOPAT instead of EBIT."),
    ("return_on_tangible_equity", "ratio", "Net income divided by tangible shareholders' equity."),
    ("effective_tax_rate", "ratio", "Income-tax expense divided by pre-tax income."),
    ("tax_burden", "ratio", "DuPont tax burden — net income divided by pre-tax income."),
    ("interest_burden", "ratio", "DuPont interest burden — pre-tax income divided by EBIT."),
    ("fcf_conversion", "ratio", "Free cash flow divided by net income."),
    ("fcf_to_ebitda", "ratio", "Free cash flow divided by EBITDA."),
    ("payout_ratio", "ratio", "Dividends paid divided by net income."),
    ("retention_ratio", "ratio", "One minus the payout ratio (earnings retained)."),
    (
        "sustainable_growth_rate",
        "ratio",
        "ROE multiplied by the retention ratio — internally fundable growth.",
    ),
]

# --- per_share (11) — TTM available ---------------------------------------
_PER_SHARE_RATIOS: list[tuple[str, str, str]] = [
    ("book_value_per_share", "USD/share", "Shareholders' equity divided by diluted shares."),
    (
        "retained_earnings_per_share",
        "USD/share",
        "Retained earnings divided by diluted shares.",
    ),
    (
        "owner_earnings_per_share",
        "USD/share",
        "Owner earnings (Buffett definition) divided by diluted shares.",
    ),
    ("fcf_per_share", "USD/share", "Free cash flow divided by diluted shares."),
    ("dividends_per_share", "USD/share", "Cash dividends declared divided by diluted shares."),
    ("sales_per_share", "USD/share", "Revenue divided by diluted shares."),
    ("cash_per_share", "USD/share", "Cash & equivalents divided by diluted shares."),
    ("ebitda_per_share", "USD/share", "EBITDA divided by diluted shares."),
    (
        "tangible_book_value_per_share",
        "USD/share",
        "Tangible shareholders' equity divided by diluted shares.",
    ),
    ("eps_basic", "USD/share", "Basic earnings per share."),
    ("eps_diluted", "USD/share", "Diluted earnings per share."),
]

# --- liquidity (5) — TTM available ----------------------------------------
_LIQUIDITY_RATIOS: list[tuple[str, str, str]] = [
    ("current_ratio", "ratio", "Current assets divided by current liabilities."),
    (
        "quick_ratio",
        "ratio",
        "Current assets excluding inventory divided by current liabilities.",
    ),
    ("cash_ratio", "ratio", "Cash & equivalents divided by current liabilities."),
    ("working_capital", "USD", "Current assets minus current liabilities."),
    ("net_working_capital", "USD", "Operating working capital, net of cash and debt."),
]

# --- efficiency (15) — TTM available --------------------------------------
_EFFICIENCY_RATIOS: list[tuple[str, str, str]] = [
    ("days_sales_outstanding", "days", "Average collection period on receivables."),
    ("days_inventory_outstanding", "days", "Average days inventory is held before sale."),
    ("days_payable_outstanding", "days", "Average days taken to pay suppliers."),
    (
        "cash_conversion_cycle",
        "days",
        "DSO + DIO − DPO — days to convert investments into cash.",
    ),
    ("cash_to_earnings", "ratio", "Operating cash flow divided by net income."),
    ("capex_to_depreciation", "ratio", "Capital expenditure divided by depreciation."),
    ("capex_to_revenue", "ratio", "Capital expenditure divided by revenue."),
    ("asset_turnover", "ratio", "Revenue divided by average total assets."),
    ("equity_multiplier", "ratio", "DuPont leverage — average assets divided by average equity."),
    (
        "defensive_interval_ratio",
        "days",
        "Days of operating expenses covered by liquid assets.",
    ),
    ("fixed_asset_turnover", "ratio", "Revenue divided by average net fixed assets."),
    ("inventory_turnover", "ratio", "Cost of revenue divided by average inventory."),
    ("receivables_turnover", "ratio", "Revenue divided by average receivables."),
    ("payables_turnover", "ratio", "Cost of revenue divided by average payables."),
    ("working_capital_turnover", "ratio", "Revenue divided by average working capital."),
]

# --- leverage (14) — TTM available ----------------------------------------
_LEVERAGE_RATIOS: list[tuple[str, str, str]] = [
    ("gross_debt", "USD", "Total interest-bearing debt (short + long term)."),
    ("net_debt", "USD", "Gross debt minus cash & equivalents."),
    ("net_debt_to_ebitda", "ratio", "Net debt divided by EBITDA."),
    ("liabilities_to_assets", "ratio", "Total liabilities divided by total assets."),
    ("debt_to_assets", "ratio", "Gross debt divided by total assets."),
    ("debt_to_equity", "ratio", "Gross debt divided by shareholders' equity."),
    ("liabilities_to_equity", "ratio", "Total liabilities divided by shareholders' equity."),
    (
        "long_term_debt_to_equity",
        "ratio",
        "Long-term debt divided by shareholders' equity.",
    ),
    ("interest_coverage", "ratio", "EBIT divided by interest expense."),
    ("debt_to_capital", "ratio", "Gross debt divided by total capital (debt + equity)."),
    (
        "long_term_debt_to_capital",
        "ratio",
        "Long-term debt divided by total capital.",
    ),
    ("ebitda_interest_coverage", "ratio", "EBITDA divided by interest expense."),
    ("cash_flow_to_debt", "ratio", "Operating cash flow divided by gross debt."),
    ("dividend_coverage", "ratio", "Net income divided by dividends paid."),
]

# --- forensic (5) — annual only -------------------------------------------
_FORENSIC_RATIOS: list[tuple[str, str, str]] = [
    ("piotroski_f_score", "score", "Piotroski F-Score (0–9) — fundamental strength."),
    ("altman_z_prime", "score", "Altman Z'-Score for private/non-listed manufacturers."),
    (
        "altman_z_double_prime",
        "score",
        "Altman Z''-Score for non-manufacturers / emerging markets.",
    ),
    ("beneish_m_score", "score", "Beneish M-Score — earnings-manipulation probability."),
    ("sloan_accruals", "ratio", "Sloan accrual ratio — accrual vs cash earnings quality."),
]

# --- growth / CAGR (60) — annual only, mechanically expanded ---------------
_GROWTH_BASE_METRICS: list[str] = [
    "revenue",
    "net_income",
    "gross_profit",
    "operating_income",
    "ebitda",
    "eps",
    "total_assets",
    "equity",
    "bvps",
    "owner_earnings",
    "fcfps",
    "dividends",
]
_GROWTH_HORIZONS_YEARS: tuple[int, ...] = (1, 3, 5, 7, 10)
_GROWTH_METRIC_LABELS: dict[str, str] = {
    "revenue": "revenue",
    "net_income": "net income",
    "gross_profit": "gross profit",
    "operating_income": "operating income",
    "ebitda": "EBITDA",
    "eps": "diluted EPS",
    "total_assets": "total assets",
    "equity": "shareholders' equity",
    "bvps": "book value per share",
    "owner_earnings": "owner earnings",
    "fcfps": "free cash flow per share",
    "dividends": "dividends",
}

# --- sector percentile ranks (22, category='rank') — annual only ----------------------------
_SECTOR_PCTILE_BASE: list[str] = [
    "roic",
    "return_on_equity",
    "return_on_assets",
    "roce",
    "net_margin",
    "operating_margin",
    "gross_profit_margin",
    "ebitda_margin",
    "fcf_margin",
    "ocf_margin",
    "fcf_conversion",
    "debt_to_equity",
    "net_debt_to_ebitda",
    "current_ratio",
    "interest_coverage",
    "asset_turnover",
    "inventory_turnover",
    "revenue_cagr_5y",
    "net_income_cagr_5y",
    "eps_cagr_5y",
    "piotroski_f_score",
    "altman_z_double_prime",
]


def _build_ratios() -> list[dict]:
    """Assemble the full ratio catalog as a flat list of dicts.

    Each row carries a stable key set::

        ratio_name, category, unit, ttm_available (bool), description

    The growth/CAGR family (12 base metrics × 5 horizons = 60) and the
    sector-percentile family (22) are expanded mechanically here rather than
    hand-listed, matching the ``ratio`` table's naming convention.

    Returns:
        The catalog rows in category order (``_RATIO_CATEGORY_ORDER``), with
        each category preserving its source declaration order.
    """
    rows: list[dict] = []

    def _add(category: str, items: list[tuple[str, str, str]]) -> None:
        ttm = category in _TTM_CATEGORIES
        for ratio_name, unit, description in items:
            rows.append(
                {
                    "ratio_name": ratio_name,
                    "category": category,
                    "unit": unit,
                    "ttm_available": ttm,
                    "description": description,
                }
            )

    _add("profitability", _PROFITABILITY_RATIOS)
    _add("per_share", _PER_SHARE_RATIOS)
    _add("liquidity", _LIQUIDITY_RATIOS)
    _add("efficiency", _EFFICIENCY_RATIOS)
    _add("leverage", _LEVERAGE_RATIOS)
    _add("forensic", _FORENSIC_RATIOS)

    # Growth / CAGR — expanded mechanically (12 × 5 = 60), annual-only.
    for metric in _GROWTH_BASE_METRICS:
        label = _GROWTH_METRIC_LABELS[metric]
        for years in _GROWTH_HORIZONS_YEARS:
            if years == 1:
                description = f"Year-over-year growth in {label} (the 1-year CAGR variant)."
            else:
                description = f"{years}-year compound annual growth rate of {label}."
            rows.append(
                {
                    "ratio_name": f"{metric}_cagr_{years}y",
                    "category": "growth",
                    "unit": "ratio",
                    "ttm_available": False,
                    "description": description,
                }
            )

    # Sector percentile ranks — expanded mechanically (22), annual-only.
    for base in _SECTOR_PCTILE_BASE:
        rows.append(
            {
                "ratio_name": f"{base}_sector_pctile",
                "category": "rank",
                "unit": "percentile",
                "ttm_available": False,
                "description": (
                    f"Percentile rank of `{base}` versus same-sector peers in the same "
                    "fiscal year, in [0, 1] (1 = best in sector)."
                ),
            }
        )

    return rows


# Built once at import; consumed by the render functions below.
RATIOS: list[dict] = _build_ratios()

# Shared narrative used across all three output surfaces so they cannot drift.
_RATIO_FY_TTM_NOTE = (
    "The `ratio` table holds BOTH annual (`fiscal_period = 'FY'`, `is_ttm = false`) "
    "and trailing-twelve-month (`fiscal_period = 'TTM'`, `is_ttm = true`) rows per "
    "company. TTM is the sum/normalization of the latest four reported quarters, "
    "dated at the entity's most recent quarter close, so it is the most-current "
    "read; FY rows are dated at fiscal-year close. TTM rows exist ONLY for the "
    "profitability, per_share, liquidity, efficiency, and leverage categories — "
    "forensic, growth (CAGR), and sector-percentile ranks are annual-only. "
    "Always filter on a single period (e.g. `WHERE is_ttm = FALSE`, or "
    "`WHERE fiscal_period = 'FY'`); a raw `read_table('ratio')` query that does NOT "
    "filter `is_ttm` / `fiscal_period` returns FY + TTM rows for the same metric and "
    "DOUBLE-COUNTS companies in a screen. The SDK SQL templates already filter by a "
    "single fiscal_period, so they are safe — this warning is for raw DuckDB queries."
)


# ---------------------------------------------------------------------------
# Manifest fetch + normalisation
# ---------------------------------------------------------------------------


def _manifest_url() -> str:
    """Return the manifest URL, honouring the VALUEIN_MANIFEST_URL override."""
    return (
        os.environ.get("VALUEIN_MANIFEST_URL", DEFAULT_MANIFEST_URL).strip() or DEFAULT_MANIFEST_URL
    )


def fetch_manifest(url: str | None = None) -> dict:
    """Fetch and parse the R2 manifest JSON.

    Args:
        url: Manifest URL. Defaults to ``VALUEIN_MANIFEST_URL`` or
            ``DEFAULT_MANIFEST_URL``.

    Returns:
        The parsed manifest as a dict.

    Raises:
        CatalogError: If the request fails or the body is not a JSON object.
    """
    target = url or _manifest_url()
    req = urllib.request.Request(target, headers={"Accept": "application/json"})
    try:
        with urllib.request.urlopen(req, timeout=_MANIFEST_TIMEOUT_SECONDS) as resp:
            raw = resp.read()
    except urllib.error.HTTPError as exc:  # pragma: no cover - network dependent
        raise CatalogError(
            f"Manifest request failed with HTTP {exc.code} for {target}: {exc.reason}"
        ) from exc
    except urllib.error.URLError as exc:  # pragma: no cover - network dependent
        raise CatalogError(f"Could not reach the manifest at {target}: {exc.reason}") from exc

    try:
        manifest = json.loads(raw)
    except json.JSONDecodeError as exc:
        raise CatalogError(f"Manifest at {target} returned invalid JSON: {exc}") from exc

    if not isinstance(manifest, dict):
        raise CatalogError(
            f"Manifest at {target} is not a JSON object (got {type(manifest).__name__})."
        )
    return manifest


def normalise_concepts(raw: object) -> list[dict]:
    """Normalise the manifest ``standard_concepts`` value into catalog rows.

    Handles BOTH manifest shapes:

      * ``dict`` entries (rich contract) — full fields are carried through.
      * ``str`` entries (legacy contract) — treated as name-only; definition,
        statement_type and category are left empty.

    Each returned row has a stable set of keys regardless of input shape::

        name, statement, category, unit, level, is_flow, definition,
        bloomberg_equivalent, factset_equivalent, gaap_ifrs_comparable

    Rows are sorted by concept name (case-insensitive).

    Args:
        raw: The raw ``manifest["standard_concepts"]`` value.

    Returns:
        A sorted list of normalised concept rows.

    Raises:
        CatalogError: If ``raw`` is missing/null/empty, or any entry is an
            unsupported type.
    """
    if raw is None:
        raise CatalogError(
            "Manifest field 'standard_concepts' is null. The manifest has not "
            "published the concept list yet — run a data-pipeline export "
            "(run_exports.py) first, then re-run this script."
        )
    if not isinstance(raw, list):
        raise CatalogError(
            f"Manifest field 'standard_concepts' must be a list (got {type(raw).__name__})."
        )
    if not raw:
        raise CatalogError(
            "Manifest field 'standard_concepts' is empty. The manifest has not "
            "published the concept list yet — run a data-pipeline export "
            "(run_exports.py) first, then re-run this script."
        )

    rows: list[dict] = []
    for entry in raw:
        if isinstance(entry, str):
            name = entry.strip()
            if not name:
                continue
            rows.append(_blank_row(name))
        elif isinstance(entry, dict):
            name = str(entry.get("standard_concept", "")).strip()
            if not name:
                # Skip malformed rich entries that carry no name.
                continue
            rows.append(
                {
                    "name": name,
                    "statement": _clean(entry.get("statement_type")),
                    "category": _clean(entry.get("category")),
                    "unit": _clean(entry.get("unit_default")),
                    "level": _clean(entry.get("level")),
                    "is_flow": entry.get("is_flow"),
                    "definition": _clean(entry.get("definition")),
                    "bloomberg_equivalent": _clean(entry.get("bloomberg_equivalent")),
                    "factset_equivalent": _clean(entry.get("factset_equivalent")),
                    "gaap_ifrs_comparable": _clean(entry.get("gaap_ifrs_comparable")),
                    # CPA review state — review_confidence 1.0 = accountant-verified
                    # (locked), 0.7 = provisional/auto-mapped pending review.
                    "reviewed": entry.get("reviewed"),
                    "reviewed_by": _clean(entry.get("reviewed_by")),
                    "reviewed_at": _clean(entry.get("reviewed_at")),
                    "review_confidence": entry.get("review_confidence"),
                }
            )
        else:
            raise CatalogError(
                "Unsupported entry in 'standard_concepts': expected str or dict, "
                f"got {type(entry).__name__}."
            )

    if not rows:
        raise CatalogError(
            "Manifest 'standard_concepts' contained no usable concept names. "
            "Run a data-pipeline export first, then re-run this script."
        )

    rows.sort(key=lambda r: r["name"].lower())
    return rows


def _blank_row(name: str) -> dict:
    """Build a name-only catalog row for a legacy ``list[str]`` entry."""
    return {
        "name": name,
        "statement": "",
        "category": "",
        "unit": "",
        "level": "",
        "is_flow": None,
        "definition": "",
        "bloomberg_equivalent": "",
        "factset_equivalent": "",
        "gaap_ifrs_comparable": "",
        "reviewed": None,
        "reviewed_by": "",
        "reviewed_at": "",
        "review_confidence": None,
    }


def _clean(value: object) -> str:
    """Coerce a manifest value to a trimmed string ('' for None)."""
    if value is None:
        return ""
    return str(value).strip()


def _statement_order(concepts: list[dict]) -> list[str]:
    """Build the section order: canonical statements first, then any others
    seen in the manifest, with the catch-all bucket always last."""
    seen: list[str] = []
    for c in concepts:
        stmt = c["statement"] or _UNCATEGORIZED
        if stmt not in seen:
            seen.append(stmt)
    ordered = [s for s in _STATEMENT_ORDER if s in seen]
    ordered += [s for s in seen if s not in _STATEMENT_ORDER and s != _UNCATEGORIZED]
    if _UNCATEGORIZED in seen:
        ordered.append(_UNCATEGORIZED)
    return ordered


# ---------------------------------------------------------------------------
# Generation
# ---------------------------------------------------------------------------


def generate(url: str | None = None) -> tuple[str, str, str]:
    """Generate all three catalog outputs from the live manifest.

    Args:
        url: Optional manifest URL override (defaults to env / built-in).

    Returns:
        ``(md_path, json_path, xlsx_path)``.

    Raises:
        CatalogError: If the manifest is unreachable or has not published the
            standardized-concept list.
    """
    source = url or _manifest_url()
    manifest = fetch_manifest(source)
    concepts = normalise_concepts(manifest.get("standard_concepts"))
    print(f"Sourced {len(concepts)} standardized concepts from {source}")
    print(f"Embedded {len(RATIOS)} financial ratios (stable IP, not manifest-sourced)")

    os.makedirs(_DOCS_DIR, exist_ok=True)
    md_path = os.path.normpath(os.path.join(_DOCS_DIR, "data_catalog.md"))
    json_path = os.path.normpath(os.path.join(_DOCS_DIR, "data_catalog.json"))
    xlsx_path = os.path.normpath(os.path.join(_DOCS_DIR, "DATA_CATALOG.xlsx"))
    _write_markdown(md_path, concepts, source)
    _write_json(json_path, concepts, source)
    _update_xlsx(xlsx_path, concepts)
    print(f"Generated {md_path}")
    print(f"Generated {json_path}")
    print(f"Updated   {xlsx_path}")
    return md_path, json_path, xlsx_path


def _ratios_by_category() -> dict[str, list[dict]]:
    """Group ``RATIOS`` by category, preserving category display order."""
    grouped: dict[str, list[dict]] = {cat: [] for cat in _RATIO_CATEGORY_ORDER}
    for row in RATIOS:
        grouped.setdefault(row["category"], []).append(row)
    return grouped


def _write_markdown(path: str, concepts: list[dict], source: str) -> None:
    today = date.today().isoformat()
    by_statement: dict[str, list[dict]] = {}
    for c in concepts:
        by_statement.setdefault(c["statement"] or _UNCATEGORIZED, []).append(c)

    lines: list[str] = [
        "# Valuein Data Catalog",
        "",
        f"> **Last updated**: {today}  ",
        f"> **Standardized concepts**: {len(concepts)}  ",
        "> **Historical coverage**: 1994 – present  ",
        "> **Coverage target**: ≥ 95% of all SEC EDGAR financial facts",
        "",
        "---",
        "",
        "## Overview",
        "",
        "The Valuein pipeline normalizes 15,000+ raw SEC EDGAR XBRL tags into a set of "
        "canonical financial concepts listed below.  Every fact in the dataset carries:",
        "",
        "- `standard_concept` — the canonical name from this catalog (or `'Other'` if unmapped)",
        "- `accuracy_score` — standardization confidence (0.0–1.0)",
        "",
        "### Accuracy Score Guide",
        "",
        "| Score | Meaning | Recommended use |",
        "|-------|---------|----------------|",
        "| 1.00 | Human-verified exact match | Any query |",
        "| 0.70–0.85 | US GAAP taxonomy rule | Any query |",
        "| 0.45–0.65 | Automated pattern match | Use with review |",
        "| 0.30–0.44 | Keyword heuristic | Research / exploratory only |",
        "| 0.00 | Unmapped (`standard_concept = 'Other'`) | Exclude from analytics |",
        "",
        "**Recommended filter for production queries:** `accuracy_score >= 0.70`",
        "",
        "---",
        "",
    ]

    for stmt in _statement_order(concepts):
        stmt_concepts = by_statement.get(stmt, [])
        if not stmt_concepts:
            continue
        lines += [f"## {stmt}", ""]
        for c in stmt_concepts:
            lines += [f"### `{c['name']}`", ""]
            meta: list[str] = []
            if c["unit"]:
                meta.append(f"**Unit:** {c['unit']}")
            if c["category"]:
                meta.append(f"**Category:** {c['category']}")
            if c["is_flow"] is not None:
                meta.append(f"**Flow:** {'yes' if c['is_flow'] else 'no'}")
            if meta:
                lines += ["  ·  ".join(meta), ""]
            if c["definition"]:
                lines += [c["definition"], ""]

    lines += _markdown_ratio_lines()

    lines += [
        "---",
        "",
        "*This catalog is generated from the live Valuein R2 manifest "
        f"(`{source}`). The underlying matching rules are proprietary — this "
        "document describes what each concept represents, not how it is detected.*",
        "",
    ]

    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def _markdown_ratio_lines() -> list[str]:
    """Render the additive ``## Financial Ratios`` Markdown section.

    Grouped by category; each ratio shows its unit + TTM availability + a short
    description.  Preceded by the FY-vs-TTM narrative so users do not write a
    double-counting raw query.
    """
    grouped = _ratios_by_category()
    ttm_count = sum(1 for r in RATIOS if r["ttm_available"])
    lines: list[str] = [
        "---",
        "",
        "## Financial Ratios",
        "",
        f"The pipeline also derives **{len(RATIOS)} financial ratios** per company "
        "into the `ratio` table (exported as `ratio.parquet`), spanning "
        "profitability, per-share, liquidity, efficiency, leverage, forensic, "
        "growth/CAGR, and sector-percentile ranks.",
        "",
        "### Annual (FY) vs Trailing-Twelve-Month (TTM)",
        "",
        _RATIO_FY_TTM_NOTE,
        "",
        "Each `ratio` row carries: `entity_id`, `ratio_name`, `category`, `value`, "
        "`unit`, `period_end`, `fiscal_year`, `fiscal_period` (`'FY'` | `'TTM'`), "
        "`is_ttm` (bool), `confidence_score`, `computed_at`. "
        f"Of the {len(RATIOS)} ratios, {ttm_count} are materialized for both FY and "
        f"TTM; the remaining {len(RATIOS) - ttm_count} are annual-only.",
        "",
    ]

    for category in _RATIO_CATEGORY_ORDER:
        rows = grouped.get(category, [])
        if not rows:
            continue
        label = _RATIO_CATEGORY_LABELS.get(category, category)
        ttm_flag = "TTM available" if category in _TTM_CATEGORIES else "annual-only (no TTM)"
        lines += [
            f"### {label}",
            "",
            f"*{len(rows)} ratios · {ttm_flag}*",
            "",
            "| Ratio | Unit | TTM | Description |",
            "|-------|------|-----|-------------|",
        ]
        for row in rows:
            ttm = "yes" if row["ttm_available"] else "no"
            lines.append(
                f"| `{row['ratio_name']}` | {row['unit']} | {ttm} | {row['description']} |"
            )
        lines.append("")

    return lines


def _write_json(path: str, concepts: list[dict], source: str) -> None:
    payload = {
        "generated": date.today().isoformat(),
        "source": source,
        "concept_count": len(concepts),
        "coverage_target": ">=95%",
        "accuracy_score_guide": {
            "1.00": "human_verified",
            "0.70-0.85": "taxonomy_rule",
            "0.45-0.65": "pattern_match",
            "0.30-0.44": "keyword_heuristic",
            "0.00": "unmapped",
        },
        "concepts": concepts,
        "ratio_count": len(RATIOS),
        "ttm_note": _RATIO_FY_TTM_NOTE,
        "ratios": RATIOS,
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)
        f.write("\n")


def _update_xlsx(path: str, concepts: list[dict]) -> None:
    """Add/replace the 'Standardized Concepts' sheet and refresh the Overview date."""
    wb = openpyxl.load_workbook(path)

    # --- Refresh generated date on the Overview sheet -------------------------
    overview = wb["1. Overview"]
    for row in overview.iter_rows():
        for cell in row:
            if cell.value == "Generated:" or (
                isinstance(cell.value, str) and cell.value.startswith("Generated:")
            ):
                # Value is in the next column
                next_cell = overview.cell(row=cell.row, column=cell.column + 1)
                next_cell.value = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
                break

    # --- Build (or replace) sheet "5. Standardized Concepts" -----------------
    sheet_name = "5. Standardized Concepts"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    # Styles matching the existing workbook
    header_fill = PatternFill("solid", fgColor="4A148C")  # deep purple
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="left", vertical="center", wrap_text=False)

    alt_fill = PatternFill("solid", fgColor="F3E5F5")  # very light purple for alternating rows
    body_font = Font(name="Calibri", size=11)
    wrap_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    top_align = Alignment(horizontal="left", vertical="top", wrap_text=False)

    # Column definitions: (header, width, attr)
    columns = [
        ("standard_concept", 28, "name"),
        ("Financial Statement", 22, "statement"),
        ("Category", 22, "category"),
        ("Unit", 14, "unit"),
        ("Review Confidence", 16, "review_confidence"),
        ("Definition", 80, "definition"),
    ]

    # Header row
    for col_idx, (header, width, _) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    # Data rows
    for row_idx, concept in enumerate(concepts, start=2):
        is_alt = row_idx % 2 == 0
        fill = alt_fill if is_alt else PatternFill(fill_type=None)

        for col_idx, (_, _, attr) in enumerate(columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=concept[attr])
            cell.font = body_font
            cell.fill = fill
            # Definition column wraps; others stay single-line
            cell.alignment = wrap_align if attr == "definition" else top_align

        ws.row_dimensions[row_idx].height = 60  # tall enough for wrapped definitions

    # Add a footer note below the data
    footer_row = len(concepts) + 3
    footer_cell = ws.cell(
        row=footer_row,
        column=1,
        value="Filter accuracy_score >= 0.70 for high-confidence analytics. "
        "standard_concept = 'Other' means the tag was not mapped.",
    )
    footer_cell.font = Font(name="Calibri", size=10, italic=True, color="757575")
    footer_cell.alignment = Alignment(horizontal="left")

    # --- Build (or replace) the additive 'Financial Ratios' sheet ------------
    _write_ratios_sheet(wb)

    wb.save(path)


def _write_ratios_sheet(wb: "openpyxl.workbook.workbook.Workbook") -> None:
    """Add/replace a 'Financial Ratios' worksheet.

    Purely additive — every existing sheet (Overview, the concept sheets, etc.)
    is preserved.  Mirrors the styling of the Standardized Concepts sheet.
    """
    sheet_name = "Financial Ratios"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    header_fill = PatternFill("solid", fgColor="4A148C")  # deep purple
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="left", vertical="center", wrap_text=False)

    alt_fill = PatternFill("solid", fgColor="F3E5F5")  # very light purple
    body_font = Font(name="Calibri", size=11)
    wrap_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    top_align = Alignment(horizontal="left", vertical="top", wrap_text=False)

    # --- Narrative banner rows (FY vs TTM) -----------------------------------
    note_cell = ws.cell(row=1, column=1, value="Financial Ratios — FY vs TTM")
    note_cell.font = Font(name="Calibri", size=12, bold=True, color="4A148C")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)

    banner = ws.cell(row=2, column=1, value=_RATIO_FY_TTM_NOTE)
    banner.font = Font(name="Calibri", size=10, italic=True, color="424242")
    banner.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
    ws.row_dimensions[2].height = 120

    header_row = 3
    data_start = header_row + 1

    columns = [
        ("ratio_name", 32, "ratio_name"),
        ("Category", 20, "category"),
        ("Unit", 14, "unit"),
        ("TTM Available", 16, "ttm_available"),
        ("Description", 80, "description"),
    ]

    for col_idx, (header, width, _) in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[header_row].height = 20
    ws.freeze_panes = f"A{data_start}"

    for offset, ratio in enumerate(RATIOS):
        row_idx = data_start + offset
        is_alt = row_idx % 2 == 0
        fill = alt_fill if is_alt else PatternFill(fill_type=None)
        for col_idx, (_, _, attr) in enumerate(columns, start=1):
            raw = ratio[attr]
            value = ("yes" if raw else "no") if attr == "ttm_available" else raw
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font
            cell.fill = fill
            cell.alignment = wrap_align if attr == "description" else top_align

    footer_row = data_start + len(RATIOS) + 1
    footer_cell = ws.cell(
        row=footer_row,
        column=1,
        value=(
            "TTM rows exist only for profitability, per_share, liquidity, efficiency, "
            "and leverage. Filter fiscal_period ('FY' | 'TTM') / is_ttm to avoid "
            "double-counting in raw queries."
        ),
    )
    footer_cell.font = Font(name="Calibri", size=10, italic=True, color="757575")
    footer_cell.alignment = Alignment(horizontal="left")


if __name__ == "__main__":
    generate()
